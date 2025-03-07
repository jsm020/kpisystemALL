from django.contrib.auth import logout
from django.shortcuts import render,redirect
from .models import *
from .forms import *
from django.db.models import Sum

from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404
from django.apps import apps
from django.db.models import Q
from django.core.exceptions import PermissionDenied
from functools import wraps
from django.contrib.auth import get_user_model
import openpyxl
from django.http import HttpResponse
from django.templatetags.static import static
import os
from io import BytesIO
from decimal import Decimal, InvalidOperation
from .helpers import check_and_delete_item
from .utils import get_user_models, calculate_progress
User = get_user_model()


def delete_item(request, model_name, pk):
    try:
        model_class = apps.get_model('mainSystem', model_name)
        status, message, _ = check_and_delete_item(model_class, pk, request.user)

        if not status:
            return render(request, 'error_template.html', {'error_message': message})
        
        # messages.success(request, message)
        return redirect('home')
    except LookupError:
        return render(request, 'error_template.html', {'error_message': f"Model {model_name} topilmadi."})



def not_superuser(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_superuser:
            # Agar superuser bo'lsa, admin sahifasiga yo'naltirish
            return redirect('/admin/')
        return view_func(request, *args, **kwargs)
    return wrapper





@login_required
@not_superuser
def all_data_view(request):
    if request.method == 'POST' and 'fvk_form_submit' in request.POST:
        form_fvk = FanVideoKontentForm(request.POST, user=request.user)
        if form_fvk.is_valid():
            try:
                maqola_instance = form_fvk.save(commit=False)
                maqola_instance.user = request.user
                maqola_instance.save()
                return redirect('home')
            except ValidationError as e:
                # Catch the validation error and display it on a new page
                error_message = e.messages[0]  # Access the actual error message
                return render(request, 'error_template.html', {'error_message': error_message})
    else:
        form_fvk = FanVideoKontentForm(user=request.user)

    if request.method == 'POST' and 'ned_form_submit' in request.POST:
        form_ned = NashrEtilganDarsliklarForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_ned.is_valid():
            maqola_instance = form_ned.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_ned = NashrEtilganDarsliklarForm(user=request.user)

        #############################################################################
        #ikkinchi bulim#

    if request.method == 'POST' and 'sws_form_submit' in request.POST:
        form_sws = ScopusWebOfScienceForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_sws.is_valid():
            maqola_instance = form_sws.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_sws = ScopusWebOfScienceForm(user=request.user)

    if request.method == 'POST' and 'oam_form_submit' in request.POST:
        form_oam = OAKJurnaliMaqolaForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_oam.is_valid():
            maqola_instance = form_oam.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_oam = OAKJurnaliMaqolaForm(user=request.user)

    if request.method == 'POST' and 'h_form_submit' in request.POST:
        form_h = HIndexForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_h.is_valid():
            maqola_instance = form_h.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_h = HIndexForm(user=request.user)

    if request.method == 'POST' and 'konf_form_submit' in request.POST:
        form_konf = KonferensiyaMaqolaForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_konf.is_valid():
            maqola_instance = form_konf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_konf = KonferensiyaMaqolaForm(user=request.user)
        
    if request.method == 'POST' and 'loyiha_form_submit' in request.POST:
        form_loyiha = LoyihalarTayyorlashForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_loyiha.is_valid():
            maqola_instance = form_loyiha.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_loyiha = LoyihalarTayyorlashForm(user=request.user)

    if request.method == 'POST' and 'loyiha_moliya_form_submit' in request.POST:
        form_loyiha_moliya = LoyihaMoliyaForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_loyiha_moliya.is_valid():
            maqola_instance = form_loyiha_moliya.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_loyiha_moliya = LoyihaMoliyaForm(user=request.user)


    if request.method == 'POST' and 'akt_form_submit' in request.POST:
        form_akt = AKTDasturlarForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_akt.is_valid():
            maqola_instance = form_akt.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_akt = AKTDasturlarForm(user=request.user)

    if request.method == 'POST' and 'tif_form_submit' in request.POST:
        form_tif = TalabaIlmiyFaoliyatiForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_tif.is_valid():
            maqola_instance = form_tif.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_tif = TalabaIlmiyFaoliyatiForm(user=request.user)

    if request.method == 'POST' and 'ttf_form_submit' in request.POST:
        form_ttf = TarbiyaTadbirlarForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_ttf.is_valid():
            maqola_instance = form_ttf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_ttf = TarbiyaTadbirlarForm(user=request.user)

    if request.method == 'POST' and 'dttf_form_submit' in request.POST:
        form_dttf = DarstanTashqariTadbirlarForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_dttf.is_valid():
            maqola_instance = form_dttf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_dttf = DarstanTashqariTadbirlarForm(user=request.user)


    if request.method == 'POST' and 'ttjtf_form_submit' in request.POST:
        form_ttjtf = TalabalarTurarJoyTadbirlarForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_ttjtf.is_valid():
            maqola_instance = form_ttjtf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_ttjtf = TalabalarTurarJoyTadbirlarForm(user=request.user)

    if request.method == 'POST' and 'oaif_form_submit' in request.POST:
        form_oaif = OtaOnalarIshlashForm(request.POST, request.FILES, user=request.user)  # request.FILES qo'shildi
        if form_oaif.is_valid():
            maqola_instance = form_oaif.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_oaif = OtaOnalarIshlashForm(user=request.user)

    if request.method == 'POST' and 'amsf_form_submit' in request.POST:
        form_amsf = AxborotMurobbiylikSoatForm(request.POST,user=request.user)  # request.FILES qo'shildi
        if form_amsf.is_valid():
            maqola_instance = form_amsf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_amsf = AxborotMurobbiylikSoatForm(user=request.user)

    if request.method == 'POST' and 'mtif_form_submit' in request.POST:
        form_mtif = MuhimTashabbuslarIshlariForm(request.POST,request.FILES,user=request.user)  # request.FILES qo'shildi
        if form_mtif.is_valid():
            maqola_instance = form_mtif.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_mtif = MuhimTashabbuslarIshlariForm(user=request.user)

    if request.method == 'POST' and 'bzbmf_form_submit' in request.POST:
        form_bzbmf = BirZiyoliBirMahallaForm(request.POST,request.FILES,user=request.user)  # request.FILES qo'shildi
        if form_bzbmf.is_valid():
            maqola_instance = form_bzbmf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_bzbmf = BirZiyoliBirMahallaForm(user=request.user)

#############
    if request.method == 'POST' and 'dyqf_form_submit' in request.POST:
        form_dyqf = DarslikYokiQollanmaForm(request.POST,request.FILES,user=request.user)  # request.FILES qo'shildi
        if form_dyqf.is_valid():
            maqola_instance = form_dyqf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_dyqf = DarslikYokiQollanmaForm(user=request.user)

    if request.method == 'POST' and 'dhf_form_submit' in request.POST:
        form_dhf = DissertationHimoyaForm(request.POST,request.FILES,user=request.user)  # request.FILES qo'shildi
        if form_dhf.is_valid():
            maqola_instance = form_dhf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_dhf = DissertationHimoyaForm(user=request.user)

    if request.method == 'POST' and 'irf_form_submit' in request.POST:
        form_irf = IlmiyRahbarlikForm(request.POST,request.FILES,user=request.user)  # request.FILES qo'shildi
        if form_irf.is_valid():
            maqola_instance = form_irf.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_irf = IlmiyRahbarlikForm(user=request.user)

    if request.method == 'POST' and 'hmof_form_submit' in request.POST:
        form_hmof = HorijdaMalakaOshirishForm(request.POST,request.FILES,user=request.user)  # request.FILES qo'shildi
        if form_hmof.is_valid():
            maqola_instance = form_hmof.save(commit=False)
            maqola_instance.user = request.user  # Foydalanuvchini qo'shish
            maqola_instance.save()
            return redirect('home')
    else:
        form_hmof = HorijdaMalakaOshirishForm(user=request.user)

    # Filtering objects based on the logged-in user
    scopus_maqolalar = ScopusWebOfScience.objects.filter(user=request.user)
    oak_maqolalar = OAKJurnaliMaqola.objects.filter(user=request.user)
    h_index = HIndex.objects.filter(user=request.user)
    konferensiya_maqolalar = KonferensiyaMaqola.objects.filter(user=request.user)
    loyihalar = LoyihalarTayyorlash.objects.filter(user=request.user)
    loyiha_moliya = LoyihaMoliya.objects.filter(user=request.user)
    AKT_dasturlar = AKTDasturlar.objects.filter(user=request.user)
    talaba_ilmiy = TalabaIlmiyFaoliyati.objects.filter(user=request.user)
    tarbiya_tadbirlar = TarbiyaTadbirlar.objects.filter(user=request.user)
    darstan_tashqari_tadbirlar = DarstanTashqariTadbirlar.objects.filter(user=request.user)
    talabalar_turar_joy = TalabalarTurarJoyTadbirlar.objects.filter(user=request.user)
    ota_onalar_ishlash = OtaOnalarIshlash.objects.filter(user=request.user)
    axborot_murobbiylik = AxborotMurobbiylikSoat.objects.filter(user=request.user)
    muhim_tashabbuslar = MuhimTashabbuslarIshlari.objects.filter(user=request.user)
    bir_ziyoli_bir_mahalla = BirZiyoliBirMahalla.objects.filter(user=request.user)
    oquv_yili_fanlar = OquvYiliFanlar.objects.filter(user=request.user)
    mustaqil_talim_topshiriqlari = MustaqilTalimTopshiriqlari.objects.filter(user=request.user)
    fan_video_kontent = FanVideoKontent.objects.filter(user=request.user)
    oqitish_sifati = OqitishSifati.objects.filter(user=request.user)
    nashr_etilgan_darsliklar = NashrEtilganDarsliklar.objects.filter(user=request.user)
    faol_interfaol_metodlar = FaolInterfaolMetodlar.objects.filter(user=request.user)
    darslik_yoki_qollanma = DarslikYokiQollanma.objects.filter(user=request.user)   
    dissertation_himoya = DissertationHimoya.objects.filter(user=request.user)
    ilmiy_rahbarlik = IlmiyRahbarlik.objects.filter(user=request.user)
    horijda_malaka_oshirish = HorijdaMalakaOshirish.objects.filter(user=request.user)
    #########3#################################3
    #Ўқув-методик фаолияти (максимал 30 балл)
    max_score = 2  
    total_score_oyf = oquv_yili_fanlar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_oyf = (total_score_oyf / max_score) * 100 if max_score > 0 else 0
    #
    max_score = 4  
    total_score_fim = faol_interfaol_metodlar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_fim = (total_score_fim / max_score) * 100 if max_score > 0 else 0
    # Maksimal ball
    max_score = 4  
    total_score_mtt = mustaqil_talim_topshiriqlari.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_mtt = (total_score_mtt / max_score) * 100 if max_score > 0 else 0
    #Maksimal ball
    max_score = 8  
    total_score_fvk = fan_video_kontent.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_fvk = (total_score_fvk / max_score) * 100 if max_score > 0 else 0
    #Maksimal ball
    max_score = 6  
    total_score_os = oqitish_sifati.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_os = (total_score_os / max_score) * 100 if max_score > 0 else 0
    #Maksimal ball
    max_score = 6  
    total_score_ned = nashr_etilgan_darsliklar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_ned = (total_score_ned / max_score) * 100 if max_score > 0 else 0
    #ikinchi bulim
    max_score = 15  
    total_score_sws = scopus_maqolalar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_sws = (total_score_sws / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 5  
    total_score_h = h_index.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_h = (total_score_h / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 5  
    total_score_oam = oak_maqolalar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_oam = (total_score_oam / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 2  
    total_score_konf = konferensiya_maqolalar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_konf = (total_score_konf / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 5  
    total_score_loyiha = loyihalar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_loyiha = (total_score_loyiha / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 6  
    total_score_loyiha_moliya = loyiha_moliya.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_loyiha_moliya = (total_score_loyiha_moliya / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 3  
    total_score_akt = AKT_dasturlar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_akt = (total_score_akt / max_score) * 100 if max_score > 0 else 0  
    #Maxsimal ball
    max_score = 5  
    total_score_ilmiy = talaba_ilmiy.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_ilmiy = (total_score_ilmiy / max_score) * 100 if max_score > 0 else 0  
    #Maxsimal ball
    max_score = 3  
    total_score_tartad = tarbiya_tadbirlar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_tartad = (total_score_tartad / max_score) * 100 if max_score > 0 else 0
    
    #Maxsimal ball
    max_score = 4  
    total_score_dtt = darstan_tashqari_tadbirlar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_dtt = (total_score_dtt / max_score) * 100 if max_score > 0 else 0
    
    #Maxsimal ball
    max_score = 3  
    total_score_ttj = talabalar_turar_joy.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_ttj = (total_score_ttj / max_score) * 100 if max_score > 0 else 0
  
    #Maxsimal ball
    max_score = 3  
    total_score_oai = ota_onalar_ishlash.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_oai = (total_score_oai / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 3  
    total_score_axborot = axborot_murobbiylik.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_axborot = (total_score_axborot / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 3  
    total_score_tashabbus = muhim_tashabbuslar.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_tashabbus = (total_score_tashabbus / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 5  
    total_score_bzbm = bir_ziyoli_bir_mahalla.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_bzbm = (total_score_bzbm / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 50 
    total_score_dyq = darslik_yoki_qollanma.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_dyq = (total_score_dyq / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 100  
    total_score_dhimoya = dissertation_himoya.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_dhimoya = (total_score_dhimoya / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 100  
    total_score_ilmiyrahbarlik = ilmiy_rahbarlik.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_ilmiyrahbarlik = (total_score_ilmiyrahbarlik / max_score) * 100 if max_score > 0 else 0
    #Maxsimal ball
    max_score = 100  
    total_score_hmo = horijda_malaka_oshirish.aggregate(Sum('score'))['score__sum'] or 0  # Agar ballar bo'lmasa, 0
    progress_percent_hmo = (total_score_hmo / max_score) * 100 if max_score > 0 else 0


    # Extracting model names using verbose_name
    scopus_model_name = ScopusWebOfScience._meta.verbose_name
    oak_model_name = OAKJurnaliMaqola._meta.verbose_name
    h_index_model_name = HIndex._meta.verbose_name
    konferensiya_model_name = KonferensiyaMaqola._meta.verbose_name
    loyihalar_model_name = LoyihalarTayyorlash._meta.verbose_name
    loyiha_moliya_model_name = LoyihaMoliya._meta.verbose_name
    AKT_dasturlar_model_name = AKTDasturlar._meta.verbose_name
    talaba_ilmiy_model_name = TalabaIlmiyFaoliyati._meta.verbose_name
    tarbiya_tadbirlar_model_name = TarbiyaTadbirlar._meta.verbose_name
    darstan_tashqari_tadbirlar_model_name = DarstanTashqariTadbirlar._meta.verbose_name
    talabalar_turar_joy_model_name = TalabalarTurarJoyTadbirlar._meta.verbose_name
    ota_onalar_model_name = OtaOnalarIshlash._meta.verbose_name
    axborot_murobbiylik_model_name = AxborotMurobbiylikSoat._meta.verbose_name
    muhim_tashabbuslar_model_name = MuhimTashabbuslarIshlari._meta.verbose_name
    bir_ziyoli_bir_mahalla_model_name = BirZiyoliBirMahalla._meta.verbose_name
    oquv_yili_fanlar_model_name = OquvYiliFanlar._meta.verbose_name
    mustaqil_talim_model_name = MustaqilTalimTopshiriqlari._meta.verbose_name
    fan_video_kontent_model_name = FanVideoKontent._meta.verbose_name
    oqitish_sifati_model_name = OqitishSifati._meta.verbose_name
    nashr_etilgan_darsliklar_model_name = NashrEtilganDarsliklar._meta.verbose_name
    faol_interfaol_metodlar_model_name = FaolInterfaolMetodlar._meta.verbose_name
    darslik_yoki_qollanma_name = DarslikYokiQollanma._meta.verbose_name
    dissertation_himoya_name = DissertationHimoya._meta.verbose_name
    ilmiy_rahbarlik_name = IlmiyRahbarlik._meta.verbose_name
    horijda_malaka_oshirish_name = HorijdaMalakaOshirish._meta.verbose_name
    

    context = {
        'form_sws':form_sws,
        'form_fvk':form_fvk,
        'form_ned':form_ned,
        'form_oam':form_oam,
        'form_h':form_h,
        'form_konf':form_konf,
        'form_loyiha':form_loyiha,
        'form_loyiha_moliya':form_loyiha_moliya,
        'form_akt':form_akt,
        'form_tif':form_tif,
        'form_ttf':form_ttf,
        'form_dttf':form_dttf,
        'form_ttjtf':form_ttjtf,
        'form_oaif':form_oaif,
        'form_amsf':form_amsf,
        'form_mtif':form_mtif,
        'form_bzbmf':form_bzbmf,
        'form_dyqf':form_dyqf,
        'form_dhf':form_dhf,
        'form_irf':form_irf,
        'form_hmof':form_hmof,
        # 'oquvyili_form':oquvyili_form,
        'total_score_oyf':total_score_oyf,
        'progress_percent_oyf':progress_percent_oyf,
        'total_score_fim':total_score_fim,
        'progress_percent_fim':progress_percent_fim, 
        'total_score_mtt':total_score_mtt,
        'progress_percent_mtt':progress_percent_mtt,
        'total_score_fvk':total_score_fvk,
        'progress_percent_fvk':progress_percent_fvk,
        'total_score_os':total_score_os,
        'progress_percent_os':progress_percent_os,
        'total_score_ned':total_score_ned,
        'progress_percent_ned':progress_percent_ned,
        'total_score_sws':total_score_sws,
        'progress_percent_sws':progress_percent_sws,
        'total_score_oam':total_score_oam,  
        'progress_percent_oam':progress_percent_oam,
        'total_score_h':total_score_h,  
        "progress_percent_h":progress_percent_h,
        'total_score_konf':total_score_konf,  
        'progress_percent_konf':progress_percent_konf,
        'total_score_loyiha':total_score_loyiha,
        'progress_percent_loyiha':progress_percent_loyiha,
        'total_score_loyiha_moliya':total_score_loyiha_moliya,
        'progress_percent_loyiha_moliya':progress_percent_loyiha_moliya,
        'total_score_akt':total_score_akt,
        'progress_percent_akt':progress_percent_akt,
        'total_score_ilmiy':total_score_ilmiy,
        'progress_percent_ilmiy':progress_percent_ilmiy,
        'total_score_tartad':total_score_tartad,
        'progress_percent_tartad':progress_percent_tartad,
        'total_score_dtt':total_score_dtt,
        'progress_percent_dtt':progress_percent_dtt, 
        'total_score_ttj':total_score_ttj,
        'progress_percent_ttj':progress_percent_ttj,
        'total_score_oai':total_score_oai,
        'progress_percent_oai':progress_percent_oai,
        'total_score_axborot':total_score_axborot,
        'progress_percent_axborot':progress_percent_axborot,
        'total_score_tashabbus':total_score_tashabbus,
        'progress_percent_tashabbus':progress_percent_tashabbus,
        'total_score_bzbm':total_score_bzbm,
        'progress_percent_bzbm':progress_percent_bzbm,
        'total_score_dyq':total_score_dyq,
        'progress_percent_dyq':progress_percent_dyq,
        'total_score_dhimoya':total_score_dhimoya,
        'progress_percent_dhimoya':progress_percent_dhimoya,
        'total_score_ilmiyrahbarlik':total_score_ilmiyrahbarlik,
        'progress_percent_ilmiyrahbarlik':progress_percent_ilmiyrahbarlik,
        'total_score_hmo':total_score_hmo,
        'progress_percent_hmo':progress_percent_hmo,
        #########################################################################
        'scopus_maqolalar': scopus_maqolalar,
        'oak_maqolalar': oak_maqolalar,
        'h_index': h_index,
        'konferensiya_maqolalar': konferensiya_maqolalar,
        'loyihalar': loyihalar,
        'loyiha_moliya': loyiha_moliya,
        'AKT_dasturlar': AKT_dasturlar,
        'talaba_ilmiy': talaba_ilmiy,
        'tarbiya_tadbirlar': tarbiya_tadbirlar,
        'darstan_tashqari_tadbirlar': darstan_tashqari_tadbirlar,
        'talabalar_turar_joy': talabalar_turar_joy,
        'ota_onalar': ota_onalar_ishlash,
        'axborot_murobbiylik': axborot_murobbiylik,
        'muhim_tashabbuslar': muhim_tashabbuslar,
        'bir_ziyoli_bir_mahalla': bir_ziyoli_bir_mahalla,
        'oquv_yili_fanlar': oquv_yili_fanlar,
        'mustaqil_talim': mustaqil_talim_topshiriqlari,
        'fan_video_kontent': fan_video_kontent,
        'oqitish_sifati': oqitish_sifati,
        'nashr_etilgan_darsliklar': nashr_etilgan_darsliklar,
        'faol_interfaol_metodlar': faol_interfaol_metodlar,
        "darslik_yoki_qollanma":darslik_yoki_qollanma,
        "dissertation_himoya":dissertation_himoya,
        "ilmiy_rahbarlik":ilmiy_rahbarlik,
        "horijda_malaka_oshirish":horijda_malaka_oshirish,
        # Add model names to the context
        'scopus_model_name': scopus_model_name,
        'oak_model_name': oak_model_name,
        'h_index_model_name': h_index_model_name,
        'konferensiya_model_name': konferensiya_model_name,
        'loyihalar_model_name': loyihalar_model_name,
        'loyiha_moliya_model_name': loyiha_moliya_model_name,
        'AKT_dasturlar_model_name': AKT_dasturlar_model_name,
        'talaba_ilmiy_model_name': talaba_ilmiy_model_name,
        'tarbiya_tadbirlar_model_name': tarbiya_tadbirlar_model_name,
        'darstan_tashqari_tadbirlar_model_name': darstan_tashqari_tadbirlar_model_name,
        'talabalar_turar_joy_model_name': talabalar_turar_joy_model_name,
        'ota_onalar_model_name': ota_onalar_model_name,
        'axborot_murobbiylik_model_name': axborot_murobbiylik_model_name,
        'muhim_tashabbuslar_model_name': muhim_tashabbuslar_model_name,
        'bir_ziyoli_bir_mahalla_model_name': bir_ziyoli_bir_mahalla_model_name,
        'oquv_yili_fanlar_model_name': oquv_yili_fanlar_model_name,
        'mustaqil_talim_model_name': mustaqil_talim_model_name,
        'fan_video_kontent_model_name': fan_video_kontent_model_name,
        'oqitish_sifati_model_name': oqitish_sifati_model_name,
        'nashr_etilgan_darsliklar_model_name': nashr_etilgan_darsliklar_model_name,
        'faol_interfaol_metodlar_model_name': faol_interfaol_metodlar_model_name,
        'darslik_yoki_qollanma_name':darslik_yoki_qollanma_name,
        'dissertation_himoya_name':dissertation_himoya_name,
        'ilmiy_rahbarlik_name':ilmiy_rahbarlik_name,
        'horijda_malaka_oshirish_name':horijda_malaka_oshirish_name,
    }

    return render(request, 'custom_page.html', context)

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        # Foydalanuvchini autentifikatsiya qilish
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Foydalanuvchini tizimga kirgizish
            login(request, user)
            
            # Agar foydalanuvchi superuser bo'lsa, admin sahifasiga yo'naltirish
            if user.is_superuser:
                return redirect('admin')  # Admin sahifasiga yo'naltirish
            else:
                return redirect('home')  # Oddiy foydalanuvchini asosiy sahifaga yo'naltirish
        else:
            # Login yoki parol xato bo'lsa xabar chiqarish
            messages.error(request, 'Login yoki parol noto‘g‘ri.')
    return render(request, 'login.html')  # Login sahifasini qayta ko'rsatish


def logout_view(request):
    logout(request)
    return redirect('login')  # Logoutdan keyin login sahifasiga yo'naltirish

def custom_error_view(request, exception=None):
    return render(request, 'error.html')





def is_admin(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and (request.user.is_superuser or request.user.is_staff):
            # User is authenticated and is either a superuser or a staff user
            return view_func(request, *args, **kwargs)
        else:
            # Render an error page if the user is not authorized
            return render(request, 'error_template.html', {'error_message': "Sizda ushbu sahifaga kirish uchun huquq yo'q."})
    return wrapper

@is_admin
def kafedralar_jadvali(request):
    User = get_user_model()
    if request.user.kafedra == "*":
        users = User.objects.filter(is_superuser=False)
    # Agar foydalanuvchining kafedrasi bo'lsa, faqat shu kafedradagi foydalanuvchilarni ko'rsatamiz.
    elif request.user.kafedra:
        users = User.objects.filter(
            is_superuser=False, 
            kafedra=request.user.kafedra
        )
    # Aks holda hech qanday foydalanuvchini ko'rsatmaymiz.
    else:
        users = User.objects.none()
    user_data = []

    # Foydalanuvchi uchun tegishli modellarni aniqlash
    model_checks = get_user_models(request.user)

    for user in users:
        total_none_count = 0

        # Modellar bo'yicha `None` qiymatlarni hisoblash
        for model in model_checks:
            total_none_count += model.objects.filter(user=user, score__isnull=True).count()

        user_data.append({
            'user': user,
            'none_count': total_none_count,
        })

    context = {'user_data': user_data}
    return render(request, 'main.html', context)
@is_admin
def view_name(request, username):
    selected_user = get_object_or_404(User, username=username)

    if selected_user.is_superuser:
        return render(request, 'error_template.html', {'error_message': "Adminlarga bu sahifaga kirish yo'q"})

    # Guruh konfiguratsiyasi
    group_config = {
        'O‘quv-metodik faoliyati (maksimal 30 ball)': {
            'models': [
                ('oquv_yili_fanlar', OquvYiliFanlar, 2),
                ('faol_interfaol_metodlar', FaolInterfaolMetodlar, 4),
                ('mustaqil_talim_topshiriqlari', MustaqilTalimTopshiriqlari, 4),
                ('fan_video_kontent', FanVideoKontent, 8),
                ('oqitish_sifati', OqitishSifati, 6),
                ('nashr_etilgan_darsliklar', NashrEtilganDarsliklar, 6),
            ]
        },
        'Manaviy-marifiy(maksimal 24 ball)': {
            'models': [
                ('tarbiya_tadbirlar', TarbiyaTadbirlar, 3),
                ('darstan_tashqari_tadbirlar', DarstanTashqariTadbirlar, 4),
                ('talabalar_turar_joy', TalabalarTurarJoyTadbirlar, 3),
                ('ota_onalar_ishlash', OtaOnalarIshlash, 3),
                ('axborot_murobbiylik', AxborotMurobbiylikSoat, 3),
                ('muhim_tashabbuslar', MuhimTashabbuslarIshlari, 3),
                ('bir_ziyoli_bir_mahalla', BirZiyoliBirMahalla, 5),
            ]
        },
        'Ilmiy faoliyati (maksimal 46 ball)': {
            'models': [
                ('scopus_maqolalar', ScopusWebOfScience, 15),
                ('oak_maqolalar', OAKJurnaliMaqola, 5),
                ('h_index', HIndex, 5),
                ('konferensiya_maqolalar', KonferensiyaMaqola, 2),
                ('loyihalar', LoyihalarTayyorlash, 5),
                ('loyiha_moliya', LoyihaMoliya, 6),
                ('AKT_dasturlar', AKTDasturlar, 3),
                ('talaba_ilmiy', TalabaIlmiyFaoliyati, 5),
                ('darslik_yoki_qollanma', DarslikYokiQollanma, 50),
                ('dissertation_himoya', DissertationHimoya, 100),
                ('ilmiy_rahbarlik', IlmiyRahbarlik, 100),
                ('horijda_malaka_oshirish', HorijdaMalakaOshirish, 100),
            ]
        }
    }

    context = {
        'scores': {},
        'name': selected_user.first_name,
        'last_name': selected_user.last_name,
    }

    # Har bir guruhni tekshirish va ma'lumotlarni qo'shish
    for group_name, config in group_config.items():
        if request.user.groups.filter(name=group_name).exists():
            for var_name, model_class, max_score in config['models']:
                objects = model_class.objects.filter(user=selected_user)
                context[var_name] = objects
                context['scores'][var_name] = calculate_progress(objects, max_score)
                context[f"{var_name}_model_name"] = model_class._meta.verbose_name

    if not context['scores']:
        return render(request, 'error_template.html', {'error_message': "Foydalanuvchi hech qaysi guruhga tegishli emas"})

    return render(request, 'submit_page.html', context)

@is_admin
def update_scoretalim_generic(request, submit_button_name, model_name, redirect_view_name, username):
    has_uquv = request.user.groups.filter(name="O‘quv-metodik faoliyati (maksimal 30 ball)").exists()
    has_manaviy = request.user.groups.filter(name="Manaviy-marifiy(maksimal 24 ball)").exists()
    has_ilmiy = request.user.groups.filter(name="Ilmiy faoliyati (maksimal 46 ball)").exists()

    # 2) Agar user har uchala guruhda ham bo‘lsa -> taqiqlash
    if has_uquv and has_manaviy and has_ilmiy:
        return render(request, 'error_template.html', {
                'error_message': "Sizga tahrirlash (baho qo‘yish) ruxsat etilmagan"
            })
    
    if request.method == 'POST' and submit_button_name in request.POST:
        item_id = request.POST.get('item_id')
        talim_sifat_xulosasi = request.POST.get('talim_sifat_xulosasi')
        talim_sifat = request.POST.get('talim_sifat')
        izoh = request.POST.get('izoh')

        # `talim_sifat` va `talim_sifat_xulosasi` ni raqamga aylantirish
        try:
            talim_sifat = Decimal(talim_sifat)
            talim_sifat_xulosasi = Decimal(talim_sifat_xulosasi)
        except (ValueError, TypeError):
            return render(request, 'error_template.html', {
                'error_message': "Talim sifati yoki Talim sifat xulosasi noto'g'ri kiritilgan."
            })

        # Model obyektini olish va mavjudligini tekshirish
        model_class = apps.get_model('mainSystem', model_name)
        item = get_object_or_404(model_class, id=item_id)

        # Agar `score` qiymati allaqachon mavjud bo'lsa, yangilashga ruxsat bermaslik
        if item.score is not None:
            return render(request, 'error_template.html', {
                'error_message': "Siz allaqachon bu ob'ekt uchun baho yoki izoh kiritgansiz."
            })

        # `talim_sifat` va `talim_sifat_xulosasi` qiymatlarini yangilash
        item.talim_sifat = talim_sifat
        item.talim_sifat_xulosasi = talim_sifat_xulosasi
        if izoh:
            item.izoh = izoh

        # Ma'lumotlarni saqlash va xatoliklarni tekshirish
        try:
            item.save()
        except ValidationError as e:
            return render(request, 'error_template.html', {
                'error_message': "Saqlashda xato: " + ', '.join(e.messages)
            })

        # Muvaffaqiyatli yangilangandan so'ng foydalanuvchini qayta yo‘naltirish
        return redirect(redirect_view_name, username=item.user.username)

    # Agar POST so'rovi bo'lmasa yoki submit_button_name mavjud bo'lmasa, qayta yo'naltirish
    return redirect(redirect_view_name, username=username)
@is_admin
def update_score_generic(request, submit_button_name, model_name, redirect_view_name, username):
    has_uquv = request.user.groups.filter(name="O‘quv-metodik faoliyati (maksimal 30 ball)").exists()
    has_manaviy = request.user.groups.filter(name="Manaviy-marifiy(maksimal 24 ball)").exists()
    has_ilmiy = request.user.groups.filter(name="Ilmiy faoliyati (maksimal 46 ball)").exists()

    # 2) Agar user har uchala guruhda ham bo‘lsa -> taqiqlash
    if has_uquv and has_manaviy and has_ilmiy:
        return render(request, 'error_template.html', {
                'error_message': "Sizga tahrirlash (baho qo‘yish) ruxsat etilmagan"
            })
    if request.method == 'POST' and submit_button_name in request.POST:
        item_id = request.POST.get('item_id')
        score = request.POST.get('score')
        izoh = request.POST.get('izoh')  # Retrieve the 'izoh' field
        
        try:
            model_class = apps.get_model('mainSystem', model_name)  # Replace with the correct app name
            item = get_object_or_404(model_class, id=item_id)
            
            # Prevent multiple updates if the score has already been set
            if item.score is not None:  # Additional check for izoh
                return render(request, 'error_template.html', {
                    'error_message': "Siz allaqachon bu ob'ekt uchun baho yoki izoh kiritgansiz."
                })
            
            # Apply updates if values are provided and not None
            if score:
                try:
                    # Convert score to Decimal, raising error if invalid
                    item.score = Decimal(score)
                except InvalidOperation:
                    return render(request, 'error_template.html', {
                        'error_message': "Noto‘g‘ri baho formati. Iltimos, raqam kiriting."
                    })

            if izoh:
                item.izoh = izoh
            
            # Save the item with validation handling
            try:
                item.save()
            except ValidationError as e:
                return render(request, 'error_template.html', {
                    'error_message': e.messages[0]
                })

            # Redirect to the view with username parameter
            return redirect(redirect_view_name, username=item.user.username)

        except LookupError:
            return HttpResponse(f"Model {model_name} does not exist.", status=500)
def update_OquvYiliFanlar_score(request,username):
    return update_score_generic(request, 'update_score_submit_oquv_yili_fanlar', 'OquvYiliFanlar', 'view_name', username)
def update_FaolInterfaolMetodlar_score(request,username):
    return update_score_generic(request, 'update_score_submit_FaolInterfaolMetodlar', 'FaolInterfaolMetodlar', 'view_name', username)
def update_MustaqilTalimTopshiriqlari_score(request,username):
    return update_score_generic(request, 'update_score_submit_MustaqilTalimTopshiriqlari', 'MustaqilTalimTopshiriqlari', 'view_name', username)
def update_FanVideoKontent_score(request,username):
    return update_score_generic(request, 'update_score_submit_FanVideoKontent', 'FanVideoKontent', 'view_name', username)
def update_OqitishSifati_score(request,username):
    return update_scoretalim_generic(request, 'update_score_submit_OqitishSifati', 'OqitishSifati', 'view_name', username)
def update_NashrEtilganDarsliklar(request,username):
    return update_score_generic(request, 'update_score_submit_NashrEtilganDarsliklar', 'NashrEtilganDarsliklar', 'view_name', username)
def update_ScopusWebOfScience(request,username):
    return update_score_generic(request, 'update_score_submit_ScopusWebOfScience', 'ScopusWebOfScience', 'view_name', username)
def update_OAKJurnaliMaqola(request,username):
    return update_score_generic(request, 'update_score_submit_OAKJurnaliMaqola', 'OAKJurnaliMaqola', 'view_name', username)
def update_HIndex(request,username):
    return update_score_generic(request, 'update_score_submit_HIndex', 'HIndex', 'view_name', username)
def update_KonferensiyaMaqola(request,username):
    return update_score_generic(request, 'update_score_submit_KonferensiyaMaqola', 'KonferensiyaMaqola', 'view_name', username)
def update_LoyihalarTayyorlash(request,username):
    return update_score_generic(request, 'update_score_submit_LoyihalarTayyorlash', 'LoyihalarTayyorlash', 'view_name', username)
def update_LoyihaMoliya(request,username):
    return update_score_generic(request, 'update_score_submit_LoyihaMoliya', 'LoyihaMoliya', 'view_name', username)
def update_AKTDasturlar(request,username):
    return update_score_generic(request, 'update_score_submit_AKTDasturlar', 'AKTDasturlar', 'view_name', username)
def update_TalabaIlmiyFaoliyati(request,username):
    return update_score_generic(request, 'update_score_submit_TalabaIlmiyFaoliyati', 'TalabaIlmiyFaoliyati', 'view_name', username)
def update_TarbiyaTadbirlar(request,username):
    return update_score_generic(request, 'update_score_submit_TarbiyaTadbirlar', 'TarbiyaTadbirlar', 'view_name', username)
def update_DarstanTashqariTadbirlar(request,username):
    return update_score_generic(request, 'update_score_submit_DarstanTashqariTadbirlar', 'DarstanTashqariTadbirlar', 'view_name', username)
def update_TalabalarTurarJoyTadbirlar(request,username):
    return update_score_generic(request, 'update_score_submit_TalabalarTurarJoyTadbirlar', 'TalabalarTurarJoyTadbirlar', 'view_name', username)
def update_OtaOnalarIshlash(request,username):
    return update_score_generic(request, 'update_score_submit_OtaOnalarIshlash', 'OtaOnalarIshlash', 'view_name', username)
def update_AxborotMurobbiylikSoat(request,username):
    return update_score_generic(request, 'update_score_submit_AxborotMurobbiylikSoat', 'AxborotMurobbiylikSoat', 'view_name', username)
def update_MuhimTashabbuslarIshlari(request,username):
    return update_score_generic(request, 'update_score_submit_MuhimTashabbuslarIshlari', 'MuhimTashabbuslarIshlari', 'view_name', username)
def update_BirZiyoliBirMahalla(request,username):
    return update_score_generic(request, 'update_score_submit_BirZiyoliBirMahalla', 'BirZiyoliBirMahalla', 'view_name', username)
def update_DarslikYokiQollanma(request,username):
    return update_score_generic(request, 'update_score_submit_DarslikYokiQollanma', 'DarslikYokiQollanma', 'view_name', username)
def update_DissertationHimoya(request,username):
    return update_score_generic(request, 'update_score_submit_DissertationHimoya', 'DissertationHimoya', 'view_name', username)
def update_IlmiyRahbarlik(request,username):
    return update_score_generic(request, 'update_score_submit_IlmiyRahbarlik', 'IlmiyRahbarlik', 'view_name', username)
def update_HorijdaMalakaOshirish(request,username):
    return update_score_generic(request, 'update_score_submit_HorijdaMalakaOshirish', 'HorijdaMalakaOshirish', 'view_name', username)


@login_required
def add_instance(request, username, model_class, redirect_view):
    """
    Generic function to add an instance of a given model for the selected user.
    
    Args:
    - request: The HTTP request object.
    - username: The username of the user to whom the instance belongs.
    - model_class: The class of the model to be created (e.g., OquvYiliFanlar).
    - redirect_view: The name of the view to redirect to after successful creation.
    """

    selected_user = get_object_or_404(User, username=username)
    
    if request.method == 'POST':
        content_type = ContentType.objects.get_for_model(model_class)
        max_score, _ = MaxScore.objects.get_or_create(
            user=selected_user,
            content_type=content_type,
            object_id=selected_user.id,
            defaults={'max_score': Decimal('2.0')}
        )
        
        # Create an instance of the specified model
        model_class.objects.create(
            user=selected_user,
            max_score_value=max_score,
            score=None,
            izoh=None
        )

        return redirect(redirect_view, username=username)

    return render(request, 'error_template.html', {'error_message': 'Invalid request method.'})

@login_required
def add_oquv_yili_fanlar(request, username):
    return add_instance(request, username, OquvYiliFanlar, 'view_name')

@login_required
def add_faol_interfaol_metodlar(request, username):
    return add_instance(request, username, FaolInterfaolMetodlar, 'view_name')

@login_required
def add_oqitish_sifati(request, username):
    return add_instance(request, username, OqitishSifati, 'view_name')

@login_required
def add_mustaqil_talim_topshiriqlari(request, username):
    return add_instance(request, username, MustaqilTalimTopshiriqlari, 'view_name')

@login_required
def add_axborot_murobbiylik_soat(request, username):
    return add_instance(request, username, AxborotMurobbiylikSoat, 'view_name')

def export_view(request):
    # Fayl shablonini yuklash
    template_path = os.path.join(os.path.dirname(__file__), '../static/templates/self.xlsx')
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    user = request.user

    # Umumiy ballni saqlash uchun o'zgaruvchi
    total_score = 0

    # A5 katak uchun
    score_A5 = round(OquvYiliFanlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['A5'] = f"{score_A5} балл"
    total_score += score_A5

    # B5 katak uchun
    score_B5 = round(FaolInterfaolMetodlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['B5'] = f"{score_B5} балл"
    total_score += score_B5

    # C5 katak uchun
    score_C5 = round(MustaqilTalimTopshiriqlari.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['C5'] = f"{score_C5} балл"
    total_score += score_C5

    # D5 katak uchun
    score_D5 = round(FanVideoKontent.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['D5'] = f"{score_D5} балл"
    total_score += score_D5

    # E5 katak uchun
    score_E5 = round(OqitishSifati.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['E5'] = f"{score_E5} балл"
    total_score += score_E5

    # F5 katak uchun (E5 ikki marta ishlatilmasligi uchun)
    score_F5 = round(NashrEtilganDarsliklar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['F5'] = f"{score_F5} балл"
    total_score += score_F5

    # H5 katak uchun
    score_H5 = round(ScopusWebOfScience.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['H5'] = f"{score_H5} балл"
    total_score += score_H5

    # I5 katak uchun
    score_I5 = round(OAKJurnaliMaqola.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['I5'] = f"{score_I5} балл"
    total_score += score_I5

    # J5 katak uchun
    score_J5 = round(HIndex.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['J5'] = f"{score_J5} балл"
    total_score += score_J5

    # K5 katak uchun
    score_K5 = round(LoyihalarTayyorlash.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['K5'] = f"{score_K5} балл"
    total_score += score_K5

    # L5 katak uchun
    score_L5 = round(LoyihaMoliya.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['L5'] = f"{score_L5} балл"
    total_score += score_L5

    # M5 katak uchun
    score_M5 = round(AKTDasturlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['M5'] = f"{score_M5} балл"
    total_score += score_M5

    # N5 katak uchun
    score_N5 = round(TalabaIlmiyFaoliyati.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['N5'] = f"{score_N5} балл"
    total_score += score_N5

    # P5 katak uchun
    score_P5 = round(TarbiyaTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['P5'] = f"{score_P5} балл"
    total_score += score_P5

    # Q5 katak uchun
    score_Q5 = round(DarstanTashqariTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['Q5'] = f"{score_Q5} балл"
    total_score += score_Q5

    # R5 katak uchun
    score_R5 = round(TalabalarTurarJoyTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['R5'] = f"{score_R5} балл"
    total_score += score_R5

    # S5 katak uchun
    score_S5 = round(OtaOnalarIshlash.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['S5'] = f"{score_S5} балл"
    total_score += score_S5

    # T5 katak uchun
    score_T5 = round(AxborotMurobbiylikSoat.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['T5'] = f"{score_T5} балл"
    total_score += score_T5

    # U5 katak uchun
    score_U5 = round(MuhimTashabbuslarIshlari.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['U5'] = f"{score_U5} балл"
    total_score += score_U5

    # V5 katak uchun
    score_V5 = round(BirZiyoliBirMahalla.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['V5'] = f"{score_V5} балл"
    total_score += score_V5

    # X5 katak uchun
    score_X5 = round(DarslikYokiQollanma.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['X5'] = f"{score_X5} балл"
    total_score += score_X5

    # Y5 katak uchun
    score_Y5 = round(DissertationHimoya.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['Y5'] = f"{score_Y5} балл"
    total_score += score_Y5

    # Z5 katak uchun
    score_Z5 = round(IlmiyRahbarlik.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['Z5'] = f"{score_Z5} балл"
    total_score += score_Z5

    # AA5 katak uchun
    score_AA5 = round(HorijdaMalakaOshirish.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
    sheet['AA5'] = f"{score_AA5} балл"
    total_score += score_AA5

    # Umumiy ballni AA6 katakka yozamiz
    sheet['AB6'] = f"{total_score} балл"

    # Javobni tayyorlash
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="user_report_{user.username}.xlsx"'

    # Faylni saqlash va qaytarish
    workbook.save(response)
    return response


def export_view_all(request):
    template_path = os.path.join(os.path.dirname(__file__), '../static/templates/all_teacher.xlsx')
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    users = CustomUser.objects.filter(is_staff=False,is_superuser=False)
    row = 5 
    id = 0
    for user in users:
        row += 1
        id+=1
        # Foydalanuvchi ma'lumotlari
        sheet[f'A{row}'] = id
        sheet[f'B{row}'] = f"{user.first_name} {user.last_name}"
        total_score = 0

        # A5 katak uchun
        score_A5 = round(OquvYiliFanlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'C{row}'] = f"{score_A5} балл"
        total_score += score_A5

        # B5 katak uchun
        score_B5 = round(FaolInterfaolMetodlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'D{row}'] = f"{score_B5} балл"
        total_score += score_B5

        # C5 katak uchun
        score_C5 = round(MustaqilTalimTopshiriqlari.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'E{row}'] = f"{score_C5} балл"
        total_score += score_C5

        # D5 katak uchun
        score_D5 = round(FanVideoKontent.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'F{row}'] = f"{score_D5} балл"
        total_score += score_D5

        # E5 katak uchun
        score_E5 = round(OqitishSifati.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'G{row}'] = f"{score_E5} балл"
        total_score += score_E5

        # F5 katak uchun (E5 ikki marta ishlatilmasligi uchun)
        score_F5 = round(NashrEtilganDarsliklar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'H{row}'] = f"{score_F5} балл"
        total_score += score_F5

        # H5 katak uchun
        score_H5 = round(ScopusWebOfScience.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'J{row}'] = f"{score_H5} балл"
        total_score += score_H5

        # I5 katak uchun
        score_I5 = round(OAKJurnaliMaqola.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'K{row}'] = f"{score_I5} балл"
        total_score += score_I5

        # J5 katak uchun
        score_J5 = round(HIndex.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'L{row}'] = f"{score_J5} балл"
        total_score += score_J5

        # K5 katak uchun
        score_K5 = round(LoyihalarTayyorlash.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'M{row}'] = f"{score_K5} балл"
        total_score += score_K5

        # L5 katak uchun
        score_L5 = round(LoyihaMoliya.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'N{row}'] = f"{score_L5} балл"
        total_score += score_L5

        # M5 katak uchun
        score_M5 = round(AKTDasturlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'O{row}'] = f"{score_M5} балл"
        total_score += score_M5

        # N5 katak uchun
        score_N5 = round(TalabaIlmiyFaoliyati.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'P{row}'] = f"{score_N5} балл"
        total_score += score_N5

        # P5 katak uchun
        score_P5 = round(TarbiyaTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'R{row}'] = f"{score_P5} балл"
        total_score += score_P5

        # Q5 katak uchun
        score_Q5 = round(DarstanTashqariTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'S{row}'] = f"{score_Q5} балл"
        total_score += score_Q5

        # R5 katak uchun
        score_R5 = round(TalabalarTurarJoyTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'T{row}'] = f"{score_R5} балл"
        total_score += score_R5

        # S5 katak uchun
        score_S5 = round(OtaOnalarIshlash.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'U{row}'] = f"{score_S5} балл"
        total_score += score_S5

        # T5 katak uchun
        score_T5 = round(AxborotMurobbiylikSoat.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'V{row}'] = f"{score_T5} балл"
        total_score += score_T5

        # U5 katak uchun
        score_U5 = round(MuhimTashabbuslarIshlari.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'W{row}'] = f"{score_U5} балл"
        total_score += score_U5

        # V5 katak uchun
        score_V5 = round(BirZiyoliBirMahalla.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'X{row}'] = f"{score_V5} балл"
        total_score += score_V5

        # X5 katak uchun
        score_X5 = round(DarslikYokiQollanma.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'Z{row}'] = f"{score_X5} балл"
        total_score += score_X5

        # Y5 katak uchun
        score_Y5 = round(DissertationHimoya.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'AA{row}'] = f"{score_Y5} балл"
        total_score += score_Y5

        # Z5 katak uchun
        score_Z5 = round(IlmiyRahbarlik.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'AB{row}'] = f"{score_Z5} балл"
        total_score += score_Z5

        # AA5 katak uchun
        score_AA5 = round(HorijdaMalakaOshirish.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'AC{row}'] = f"{score_AA5} балл"
        total_score += score_AA5

        # Umumiy ballni AA6 katakka yozamiz
        sheet[f'AD{row}'] = f"{total_score} балл"
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="barcha_ustozlar.xlsx"'

    # Faylni saqlash va qaytarish
    workbook.save(response)
    return response

def export_view_uquv_bulim(request):
    template_path = os.path.join(os.path.dirname(__file__), '../static/templates/uquv_bulim.xlsx')
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    users = CustomUser.objects.filter(is_staff=False,is_superuser=False)
    row = 5 
    id = 0
    for user in users:
        row += 1
        id+=1
        # Foydalanuvchi ma'lumotlari
        sheet[f'A{row}'] = id        
        sheet[f'B{row}'] = f"{user.first_name} {user.last_name}"
        total_score = 0

        # A5 katak uchun
        score_A5 = round(OquvYiliFanlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'C{row}'] = f"{score_A5} балл"
        total_score += score_A5

        # B5 katak uchun
        score_B5 = round(FaolInterfaolMetodlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'D{row}'] = f"{score_B5} балл"
        total_score += score_B5

        # C5 katak uchun
        score_C5 = round(MustaqilTalimTopshiriqlari.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'E{row}'] = f"{score_C5} балл"
        total_score += score_C5

        # D5 katak uchun
        score_D5 = round(FanVideoKontent.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'F{row}'] = f"{score_D5} балл"
        total_score += score_D5

        # E5 katak uchun
        score_E5 = round(OqitishSifati.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'G{row}'] = f"{score_E5} балл"
        total_score += score_E5

        # F5 katak uchun (E5 ikki marta ishlatilmasligi uchun)
        score_F5 = round(NashrEtilganDarsliklar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'H{row}'] = f"{score_F5} балл"
        total_score += score_F5

        # Umumiy ballni AA6 katakka yozamiz
        sheet[f'I{row}'] = f"{total_score} балл"
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="uquv_bulim.xlsx"'

    # Faylni saqlash va qaytarish
    workbook.save(response)
    return response

def export_view_ilmiy_bulim(request):
    template_path = os.path.join(os.path.dirname(__file__), '../static/templates/ilmiy_bulim.xlsx')
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    users = CustomUser.objects.filter(is_staff=False,is_superuser=False)
    row = 5
    id = 0
    for user in users:
        row += 1
        id+=1
        # Foydalanuvchi ma'lumotlari
        sheet[f'A{row}'] = id
        sheet[f'B{row}'] = f"{user.first_name} {user.last_name}"
        total_score = 0

        # H5 katak uchun
        score_H5 = round(ScopusWebOfScience.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'C{row}'] = f"{score_H5} балл"
        total_score += score_H5

        # I5 katak uchun
        score_I5 = round(OAKJurnaliMaqola.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'D{row}'] = f"{score_I5} балл"
        total_score += score_I5

        # J5 katak uchun
        score_J5 = round(HIndex.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'E{row}'] = f"{score_J5} балл"
        total_score += score_J5

        # K5 katak uchun
        score_K5 = round(LoyihalarTayyorlash.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'F{row}'] = f"{score_K5} балл"
        total_score += score_K5

        # L5 katak uchun
        score_L5 = round(LoyihaMoliya.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'G{row}'] = f"{score_L5} балл"
        total_score += score_L5

        # M5 katak uchun
        score_M5 = round(AKTDasturlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'H{row}'] = f"{score_M5} балл"
        total_score += score_M5

        # N5 katak uchun
        score_N5 = round(TalabaIlmiyFaoliyati.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'I{row}'] = f"{score_N5} балл"
        total_score += score_N5
        
        score_X5 = round(DarslikYokiQollanma.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'K{row}'] = f"{score_X5} балл"
        total_score += score_X5

        # Y5 katak uchun
        score_Y5 = round(DissertationHimoya.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'L{row}'] = f"{score_Y5} балл"
        total_score += score_Y5

        # Z5 katak uchun
        score_Z5 = round(IlmiyRahbarlik.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'M{row}'] = f"{score_Z5} балл"
        total_score += score_Z5

        # AA5 katak uchun
        score_AA5 = round(HorijdaMalakaOshirish.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'N{row}'] = f"{score_AA5} балл"
        total_score += score_AA5

        sheet[f'O{row}'] = f"{total_score} балл"
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ilmiy_bulim.xlsx"'

    # Faylni saqlash va qaytarish
    workbook.save(response)
    return response

def export_view_manaviy_bulim(request):
    template_path = os.path.join(os.path.dirname(__file__), '../static/templates/manaviy_marifiy.xlsx')
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    users = CustomUser.objects.filter(is_staff=False,is_superuser=False)
    row = 5 
    id = 0
    for user in users:
        row += 1
        id+=1
        # Foydalanuvchi ma'lumotlari
        sheet[f'A{row}'] = id
        sheet[f'B{row}'] = f"{user.first_name} {user.last_name}"
        total_score = 0

        score_P5 = round(TarbiyaTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'C{row}'] = f"{score_P5} балл"
        total_score += score_P5

        # Q5 katak uchun
        score_Q5 = round(DarstanTashqariTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'D{row}'] = f"{score_Q5} балл"
        total_score += score_Q5

        # R5 katak uchun
        score_R5 = round(TalabalarTurarJoyTadbirlar.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'E{row}'] = f"{score_R5} балл"
        total_score += score_R5

        # S5 katak uchun
        score_S5 = round(OtaOnalarIshlash.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'F{row}'] = f"{score_S5} балл"
        total_score += score_S5

        # T5 katak uchun
        score_T5 = round(AxborotMurobbiylikSoat.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'G{row}'] = f"{score_T5} балл"
        total_score += score_T5

        # U5 katak uchun
        score_U5 = round(MuhimTashabbuslarIshlari.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'H{row}'] = f"{score_U5} балл"
        total_score += score_U5

        # V5 katak uchun
        score_V5 = round(BirZiyoliBirMahalla.objects.filter(user=user).aggregate(total=Sum('score'))['total'] or 0, 2)
        sheet[f'I{row}'] = f"{score_V5} балл"
        total_score += score_V5


        # Umumiy ballni AA6 katakka yozamiz
        sheet[f'J{row}'] = f"{total_score} балл"
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="manaviy_marifiy.xlsx"'

    # Faylni saqlash va qaytarish
    workbook.save(response)
    return response

